#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""数据导出模块

支持导出格式：JSON, CSV, Excel, Markdown
"""

import os
import json
import csv

from utils.time_utils import now_bj
from storage.database import db_get_all_for_export
from storage.models import NewsItem


def export_to_json(output_path: str, start_date=None, end_date=None) -> int:
    """导出新闻为 JSON 文件"""
    news = db_get_all_for_export(start_date, end_date)
    news_dicts = [n.to_dict() for n in news]
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(news_dicts, f, ensure_ascii=False, indent=2)
    return len(news)


def export_to_csv(output_path: str, start_date=None, end_date=None) -> int:
    """导出新闻为 CSV 文件（支持 Excel 直接打开）"""
    news = db_get_all_for_export(start_date, end_date)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["标题", "链接", "来源", "发布时间", "时间戳", "简介", "分类", "关键词", "股票代码"])
        for n in news:
            writer.writerow([
                n.title, n.url, n.source, n.publish_time, n.publish_ts, n.intro,
                n.category, ", ".join(n.keywords), ", ".join(n.stocks)
            ])
    return len(news)


def export_to_markdown(output_path: str, start_date=None, end_date=None,
                        group_by_date: bool = True, group_by_source: bool = False) -> int:
    """导出新闻为 Markdown 文件

    Args:
        output_path: 输出文件路径
        start_date: 起始日期
        end_date: 截止日期
        group_by_date: 是否按日期分组
        group_by_source: 是否按来源分组

    Returns:
        导出的新闻数量
    """
    news = db_get_all_for_export(start_date, end_date)
    if not news:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("# 财经新闻导出\n\n暂无数据\n")
        return 0

    lines = []
    timestamp = now_bj().strftime("%Y-%m-%d %H:%M:%S")
    lines.append("# 财经新闻导出")
    lines.append("")
    lines.append(f"> 导出时间：{timestamp}")
    lines.append(f"> 新闻数量：{len(news)} 条")
    if start_date or end_date:
        lines.append(f"> 时间范围：{start_date or '不限'} ~ {end_date or '不限'}")
    lines.append("")

    if group_by_date:
        # 按日期分组
        from collections import defaultdict
        date_groups = defaultdict(list)
        for n in news:
            date_key = n.publish_time[:10] if n.publish_time else "未知日期"
            date_groups[date_key].append(n)

        for date in sorted(date_groups.keys(), reverse=True):
            day_news = date_groups[date]
            lines.append(f"## {date} ({len(day_news)} 条)")
            lines.append("")

            if group_by_source:
                source_groups = defaultdict(list)
                for n in day_news:
                    source_groups[n.source].append(n)
                for source in source_groups:
                    lines.append(f"### {source}")
                    lines.append("")
                    for n in source_groups[source]:
                        lines.append(f"- [{n.title}]({n.url})")
                        lines.append(f"  > {n.publish_time} | {n.source}")
                        if n.intro:
                            intro = n.intro.strip()[:100] + "..." if len(n.intro) > 100 else n.intro.strip()
                            lines.append(f"  > {intro}")
                        lines.append("")
            else:
                for n in day_news:
                    lines.append(f"### [{n.title}]({n.url})")
                    lines.append("")
                    lines.append(f"- **来源**：{n.source}")
                    lines.append(f"- **时间**：{n.publish_time}")
                    if n.category:
                        lines.append(f"- **分类**：{n.category}")
                    if n.keywords:
                        lines.append(f"- **关键词**：{', '.join(n.keywords)}")
                    if n.stocks:
                        lines.append(f"- **涉及股票**：{', '.join(n.stocks)}")
                    if n.intro:
                        intro = n.intro.strip()
                        lines.append("")
                        lines.append(f"> {intro}")
                    lines.append("")
    else:
        # 平铺列表
        for n in news:
            lines.append(f"### [{n.title}]({n.url})")
            lines.append("")
            lines.append(f"- **来源**：{n.source}")
            lines.append(f"- **时间**：{n.publish_time}")
            if n.category:
                lines.append(f"- **分类**：{n.category}")
            if n.keywords:
                lines.append(f"- **关键词**：{', '.join(n.keywords)}")
            if n.stocks:
                lines.append(f"- **涉及股票**：{', '.join(n.stocks)}")
            if n.intro:
                intro = n.intro.strip()
                lines.append("")
                lines.append(f"> {intro}")
            lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return len(news)


def export_to_excel(output_path: str, start_date=None, end_date=None) -> int:
    """导出新闻为 Excel 文件（需要 openpyxl）

    如果未安装 openpyxl，将降级为 CSV 导出。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # 降级为 CSV
        csv_path = output_path.replace(".xlsx", ".csv").replace(".xls", ".csv")
        return export_to_csv(csv_path, start_date, end_date)

    news = db_get_all_for_export(start_date, end_date)
    if not news:
        wb = Workbook()
        ws = wb.active
        ws.title = "新闻列表"
        ws["A1"] = "暂无数据"
        wb.save(output_path)
        return 0

    wb = Workbook()

    # Sheet1: 新闻列表
    ws1 = wb.active
    ws1.title = "新闻列表"

    headers = ["序号", "标题", "链接", "来源", "发布时间", "分类", "关键词", "涉及股票", "简介"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, n in enumerate(news, 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=n.title)
        ws1.cell(row=row, column=3, value=n.url)
        ws1.cell(row=row, column=4, value=n.source)
        ws1.cell(row=row, column=5, value=n.publish_time)
        ws1.cell(row=row, column=6, value=n.category)
        ws1.cell(row=row, column=7, value=", ".join(n.keywords))
        ws1.cell(row=row, column=8, value=", ".join(n.stocks))
        ws1.cell(row=row, column=9, value=n.intro)

    # 调整列宽
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["H"].width = 20
    ws1.column_dimensions["I"].width = 60

    # Sheet2: 来源统计
    ws2 = wb.create_sheet("来源统计")
    from collections import Counter
    source_counts = Counter(n.source for n in news)
    ws2["A1"] = "来源"
    ws2["B1"] = "数量"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for i, (source, count) in enumerate(source_counts.most_common(), 1):
        ws2.cell(row=i + 1, column=1, value=source)
        ws2.cell(row=i + 1, column=2, value=count)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    return len(news)


def get_default_export_path(fmt: str) -> str:
    """生成默认导出文件路径"""
    timestamp = now_bj().strftime("%Y%m%d_%H%M%S")
    if fmt == "excel":
        ext = "xlsx"
    else:
        ext = fmt
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        f"news_export_{timestamp}.{ext}"
    )

