#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""数据导出模块

支持导出格式：JSON, CSV, Excel, Markdown
"""

import csv
import json
import os
from collections import Counter, defaultdict
from typing import List, Optional

from finfeed.utils.time_utils import now_bj

from .database import db_get_all_for_export
from .models import NewsItem


def _get_news(news: Optional[List[NewsItem]] = None, start_date=None, end_date=None) -> List[NewsItem]:
    """获取新闻列表：优先使用传入的列表，否则从数据库查询"""
    if news is not None:
        return news
    return db_get_all_for_export(start_date, end_date)


def export_to_json(output_path: str, news: Optional[List[NewsItem]] = None,
                   start_date=None, end_date=None) -> int:
    """导出新闻为 JSON 文件"""
    items = _get_news(news, start_date, end_date)
    news_dicts = [n.to_dict() for n in items]
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(news_dicts, f, ensure_ascii=False, indent=2)
    return len(items)


def export_to_csv(output_path: str, news: Optional[List[NewsItem]] = None,
                  start_date=None, end_date=None) -> int:
    """导出新闻为 CSV 文件（支持 Excel 直接打开）"""
    items = _get_news(news, start_date, end_date)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["标题", "链接", "来源", "分类", "发布时间", "时间戳", "简介",
                         "情绪", "重要性", "关键词", "关联股票", "已收藏"])
        for n in items:
            writer.writerow([
                n.title, n.url, n.source, n.category, n.publish_time, n.publish_ts,
                n.intro, n.sentiment, f"{n.importance:.2f}",
                ", ".join(n.keywords) if n.keywords else "",
                ", ".join(n.stocks) if n.stocks else "",
                "是" if n.is_favorite else "否",
            ])
    return len(items)


def export_to_markdown(output_path: str, news: Optional[List[NewsItem]] = None,
                       start_date=None, end_date=None,
                       group_by_date: bool = True, group_by_source: bool = False) -> int:
    """导出新闻为 Markdown 文件"""
    items = _get_news(news, start_date, end_date)
    if not items:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("# 财经新闻导出\n\n暂无数据\n")
        return 0

    lines = []
    timestamp = now_bj().strftime("%Y-%m-%d %H:%M:%S")
    lines.append("# 财经新闻导出")
    lines.append("")
    lines.append(f"> 导出时间：{timestamp}")
    lines.append(f"> 新闻数量：{len(items)} 条")
    if start_date or end_date:
        lines.append(f"> 时间范围：{start_date or '不限'} ~ {end_date or '不限'}")
    lines.append("")

    if group_by_date:
        date_groups = defaultdict(list)
        for n in items:
            date_key = n.publish_time[:10] if n.publish_time else "未知日期"
            date_groups[date_key].append(n)

        for date in sorted(date_groups.keys(), reverse=True):
            day_news = date_groups[date]
            lines.append(f"## {date} ({len(day_news)} 条)")
            lines.append("")

            if group_by_source:
                source_groups = defaultdict(list)
                for n in day_news:
                    source_groups[n.source].append(n)
                for source in source_groups:
                    lines.append(f"### {source}")
                    lines.append("")
                    for n in source_groups[source]:
                        lines.append(f"- [{n.title}]({n.url})")
                        lines.append(f"  > {n.publish_time} | {n.source}")
                        if n.intro:
                            intro = n.intro.strip()[:100] + "..." if len(n.intro) > 100 else n.intro.strip()
                            lines.append(f"  > {intro}")
                        lines.append("")
            else:
                for n in day_news:
                    lines.append(f"### [{n.title}]({n.url})")
                    lines.append("")
                    lines.append(f"- **来源**：{n.source}")
                    lines.append(f"- **时间**：{n.publish_time}")
                    if n.category:
                        lines.append(f"- **分类**：{n.category}")
                    if n.sentiment:
                        lines.append(f"- **情绪**：{n.sentiment}")
                    if n.keywords:
                        lines.append(f"- **关键词**：{', '.join(n.keywords)}")
                    if n.stocks:
                        lines.append(f"- **涉及股票**：{', '.join(n.stocks)}")
                    if n.intro:
                        intro = n.intro.strip()
                        lines.append("")
                        lines.append(f"> {intro}")
                    lines.append("")
    else:
        for n in items:
            lines.append(f"### [{n.title}]({n.url})")
            lines.append("")
            lines.append(f"- **来源**：{n.source}")
            lines.append(f"- **时间**：{n.publish_time}")
            if n.category:
                lines.append(f"- **分类**：{n.category}")
            if n.sentiment:
                lines.append(f"- **情绪**：{n.sentiment}")
            if n.keywords:
                lines.append(f"- **关键词**：{', '.join(n.keywords)}")
            if n.stocks:
                lines.append(f"- **涉及股票**：{', '.join(n.stocks)}")
            if n.intro:
                intro = n.intro.strip()
                lines.append("")
                lines.append(f"> {intro}")
            lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return len(items)


def export_to_excel(output_path: str, news: Optional[List[NewsItem]] = None,
                    start_date=None, end_date=None) -> int:
    """导出新闻为 Excel 文件（需要 openpyxl）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        csv_path = output_path.replace(".xlsx", ".csv").replace(".xls", ".csv")
        return export_to_csv(csv_path, news, start_date, end_date)

    items = _get_news(news, start_date, end_date)
    if not items:
        wb = Workbook()
        ws = wb.active
        ws.title = "新闻列表"
        ws["A1"] = "暂无数据"
        wb.save(output_path)
        return 0

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "新闻列表"

    headers = ["序号", "标题", "链接", "来源", "分类", "发布时间", "情绪", "重要性",
               "关键词", "涉及股票", "简介", "收藏"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, n in enumerate(items, 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=n.title)
        ws1.cell(row=row, column=3, value=n.url)
        ws1.cell(row=row, column=4, value=n.source)
        ws1.cell(row=row, column=5, value=n.category)
        ws1.cell(row=row, column=6, value=n.publish_time)
        ws1.cell(row=row, column=7, value=n.sentiment)
        ws1.cell(row=row, column=8, value=round(n.importance, 2))
        ws1.cell(row=row, column=9, value=", ".join(n.keywords))
        ws1.cell(row=row, column=10, value=", ".join(n.stocks))
        ws1.cell(row=row, column=11, value=n.intro)
        ws1.cell(row=row, column=12, value="是" if n.is_favorite else "否")

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 8
    ws1.column_dimensions["H"].width = 8
    ws1.column_dimensions["I"].width = 25
    ws1.column_dimensions["J"].width = 15
    ws1.column_dimensions["K"].width = 60
    ws1.column_dimensions["L"].width = 6

    ws2 = wb.create_sheet("来源统计")
    source_counts = Counter(n.source for n in items)
    ws2["A1"] = "来源"
    ws2["B1"] = "数量"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for i, (source, count) in enumerate(source_counts.most_common(), 1):
        ws2.cell(row=i + 1, column=1, value=source)
        ws2.cell(row=i + 1, column=2, value=count)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    return len(items)


def get_default_export_path(fmt: str) -> str:
    """生成默认导出文件路径"""
    timestamp = now_bj().strftime("%Y%m%d_%H%M%S")
    if fmt == "excel":
        ext = "xlsx"
    else:
        ext = fmt
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        f"news_export_{timestamp}.{ext}"
    )
